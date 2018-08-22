import numpy as np
import openpyxl

algorithms = ["old_cosma", "scalapack", "cyclops", "cosma"]
n_nodes = [4, 7, 8, 13, 16, 25, 27, 32, 37, 61, 64, 81, 93, 128, 201, 216, 256, 333, 473, 512]

experiments = ["square_strong", "square_weak_p0", "square_weak_p1", "thin_strong", "thin_weak_p0", "thin_weak_p1"]
experiments_short = ["square_strong", "square_weak_p0", "square_weak_p1", "thin_weak_p0", "thin_weak_p1"]

start_row = [8, 15, 22, 8, 15, 22]
start_col = [4, 4, 4, 27, 27, 27]

label_row = 5
square_col = 7

thin_mn_col = 29
thin_k_col = 32

book = openpyxl.load_workbook("experiments.xlsx")
sheet = book.active

for nodes_i, nodes in enumerate(n_nodes):
    for alg_i, alg in enumerate(algorithms):
        file_name = alg + "_" + str(nodes) + ".txt"
        curr_file = open(file_name, "r")

        lines = curr_file.readlines()
        n_experiments = len(lines)

        for line_i, line in enumerate(lines):
            curr_line_i = line_i

            if line_i > 2:
                curr_line_i = curr_line_i + 1

            row = start_row[line_i]
            col = start_col[line_i]

            row += alg_i
            col += nodes_i

            if (len(line.split()) != 6 or line.split()[0] == "not"):
                sheet.cell(row=row, column=col).value = "-"
            else:
                (nodes_str, m_str, n_str, k_str, mem_limit_str, time_str) = tuple(line.split())
                nodes = int(nodes_str)
                m = int(m_str)
                n = int(n_str)
                k = int(k_str)
                time = int(time_str)

                # output time in [ms]
                sheet.cell(row=row, column=col).value = time

                # output the problem sizes like m, n, k
                if alg == "cosma":
                    if line_i == 0:
                        sheet.cell(row=label_row, column=square_col).value = m
                    elif line_i == 3:
                        sheet.cell(row=label_row, column=thin_mn_col).value = m
                        sheet.cell(row=label_row, column=thin_k_col).value = k
                    elif line_i == 1 or line_i == 2:
                        sheet.cell(row=row - 4, column=col).value = m
                    else:
                        sheet.cell(row=row - 5, column=col).value = m
                        sheet.cell(row=row - 4, column=col).value = k


        curr_file.close()

book.save("results.xlsx")


